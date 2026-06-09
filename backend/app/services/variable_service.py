"""变量业务逻辑 — 保存、校验、Excel 导入导出、AI 去重。"""

from __future__ import annotations

import io
import re
from datetime import UTC, datetime
from typing import Any

from fastapi import HTTPException, UploadFile, status
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import AILog, Variable
from app.services import ai_service
from app.services.ai_guardrail import (
    cross_validate_dedup_suggestions,
    cross_validate_issues,
)
from app.services.project_service import get_project
from app.services.variable_registry import VALIDATION_RULES, get_merged_registry

EXCEL_HEADERS = ("变量标识(key)", "变量名称(label)", "值(value)")
_BASE_KEY_PATTERN = re.compile(r"^(.+)_(\d+)$")


def _base_key(key: str) -> str:
    match = _BASE_KEY_PATTERN.match(key)
    return match.group(1) if match else key


def _normalize_dt(dt: datetime | None) -> datetime | None:
    if dt is None:
        return None
    return dt.replace(tzinfo=None) if dt.tzinfo else dt


def _variable_to_dict(var: Variable) -> dict[str, Any]:
    return {
        "key": var.key,
        "label": var.label,
        "value": var.value or "",
        "category": var.category,
        "data_type": var.data_type,
        "required": var.required,
        "is_multiple": var.is_multiple,
        "sort_order": var.sort_order,
        "source_template_ids": var.source_template_ids or [],
        "is_merged": var.is_merged,
        "merged_from_keys": var.merged_from_keys,
        "updated_at": var.updated_at,
    }


def list_variables(db: Session, project_id: int) -> list[dict[str, Any]]:
    get_project(db, project_id)
    stmt = (
        select(Variable)
        .where(Variable.project_id == project_id)
        .order_by(Variable.sort_order, Variable.category, Variable.key)
    )
    return [_variable_to_dict(v) for v in db.scalars(stmt).all()]


def _find_variable(db: Session, project_id: int, key: str) -> Variable | None:
    return db.scalar(
        select(Variable).where(Variable.project_id == project_id, Variable.key == key)
    )


def validate_value(key: str, value: str, data_type: str) -> str | None:
    """基础正则校验，返回错误消息或 None。"""
    if not value and not value.strip():
        return None
    base = _base_key(key)
    pattern = VALIDATION_RULES.get(base) or VALIDATION_RULES.get(key)
    if pattern and not re.match(pattern, value.strip()):
        return f"变量 {key} 的值不符合格式要求"
    return None


def _try_create_multiple_row(db: Session, project_id: int, key: str) -> Variable | None:
    """为 multiple 变量动态新增行（如 handling_lawyer_3）。"""
    match = _BASE_KEY_PATTERN.match(key)
    if not match:
        return None

    base = match.group(1)
    siblings = _find_variables_by_base(db, project_id, base)
    if not siblings:
        return None

    template = siblings[0]
    var = Variable(
        project_id=project_id,
        key=key,
        label=template.label,
        value="",
        data_type=template.data_type,
        category=template.category,
        is_multiple=True,
        required=template.required,
        sort_order=template.sort_order,
        source_template_ids=list(template.source_template_ids or []),
    )
    db.add(var)
    db.flush()
    return var


def save_variables(
    db: Session,
    project_id: int,
    items: list[dict[str, Any]],
) -> dict[str, Any]:
    """逐行乐观锁保存变量值，部分成功策略。"""
    get_project(db, project_id)
    success: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    any_written = False
    for index, item in enumerate(items, start=1):
        key = item.get("key")
        value = item.get("value", "")
        client_updated_at = item.get("updated_at")

        if not key:
            errors.append({"row": index, "key": None, "message": "缺少变量 key"})
            continue

        var = _find_variable(db, project_id, key)
        if var is None:
            var = _try_create_multiple_row(db, project_id, key)
        if var is None:
            errors.append({"row": index, "key": key, "message": f"变量 {key} 不存在"})
            continue

        if client_updated_at is not None and client_updated_at != "":
            if isinstance(client_updated_at, str):
                client_updated_at = datetime.fromisoformat(
                    client_updated_at.replace("Z", "+00:00")
                )
            client_dt = _normalize_dt(client_updated_at)
            db_dt = _normalize_dt(var.updated_at)
            if client_dt != db_dt:
                errors.append(
                    {
                        "row": index,
                        "key": key,
                        "message": "数据已被其他操作修改，请刷新后重试",
                    }
                )
                continue

        fmt_error = validate_value(key, str(value), var.data_type)
        if fmt_error:
            errors.append({"row": index, "key": key, "message": fmt_error})
            continue

        var.value = str(value) if value is not None else ""
        var.updated_at = datetime.now(UTC).replace(tzinfo=None)
        any_written = True
        success.append({"key": key, "value": var.value})

    if any_written:
        db.commit()
        for row in success:
            var = _find_variable(db, project_id, row["key"])
            if var:
                row["updated_at"] = var.updated_at.isoformat()

    total = len(items)
    return {
        "success": success,
        "errors": errors,
        "summary": {
            "total": total,
            "succeeded": len(success),
            "failed": len(errors),
        },
    }


def get_alias_dedup_suggestions(variables: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """基于注册表 aliases 的规则去重建议。"""
    registry = get_merged_registry()
    suggestions: list[dict[str, Any]] = []
    seen_pairs: set[tuple[str, str]] = set()

    key_to_var = {v["key"]: v for v in variables}
    label_to_keys: dict[str, list[str]] = {}
    for var in variables:
        label_to_keys.setdefault(var["label"], []).append(var["key"])

    for var in variables:
        base = _base_key(var["key"])
        defn = registry.get(base) or registry.get(var["key"])
        if not defn:
            continue
        aliases = defn.get("aliases") or []
        candidate_keys = {base}
        for alias in aliases:
            for key in label_to_keys.get(alias, []):
                candidate_keys.add(_base_key(key))

        for other_base in candidate_keys:
            if other_base == base:
                continue
            pair = tuple(sorted([base, other_base]))
            if pair in seen_pairs:
                continue
            other_keys = [k for k in key_to_var if _base_key(k) == other_base]
            if not other_keys:
                continue
            seen_pairs.add(pair)
            suggestions.append(
                {
                    "keep_key": base,
                    "merge_keys": [other_base],
                    "reason": f"注册表别名匹配：{base} 与 {other_base} 语义相同",
                    "confidence": 1.0,
                    "source": "alias",
                }
            )
    return suggestions


async def ai_dedup_suggestions(db: Session, project_id: int) -> dict[str, Any]:
    """AI 变量去重建议（含别名规则建议）。"""
    variables = list_variables(db, project_id)
    alias_suggestions = get_alias_dedup_suggestions(variables)

    ai_used = True
    ai_suggestions: list[dict[str, Any]] = []
    message: str | None = None
    try:
        payload = [
            {
                "key": v["key"],
                "label": v["label"],
                "category": v["category"],
                "value": v.get("value", ""),
            }
            for v in variables
        ]
        result = await ai_service.suggest_variable_dedup(payload)
        ai_suggestions = result.get("suggestions", [])
        ai_suggestions, alias_suggestions = cross_validate_dedup_suggestions(
            ai_suggestions, alias_suggestions
        )
        _log_ai_call(
            db,
            project_id,
            ai_type="variable_dedup",
            prompt=str(payload),
            response=str(result),
            validation_result=result,
        )
    except ai_service.AIServiceUnavailableError as exc:
        ai_used = False
        message = str(exc)

    return {
        "alias_suggestions": alias_suggestions,
        "ai_suggestions": ai_suggestions,
        "ai_used": ai_used,
        "message": message,
    }


def apply_dedup_suggestions(
    db: Session,
    project_id: int,
    suggestions: list[dict[str, Any]],
) -> dict[str, Any]:
    """应用去重合并建议。"""
    get_project(db, project_id)
    merged_count = 0

    for suggestion in suggestions:
        keep_key = suggestion["keep_key"]
        merge_keys = suggestion.get("merge_keys") or []
        keep_var = _find_primary_variable(db, project_id, keep_key)
        if keep_var is None:
            continue

        merged_from = list(keep_var.merged_from_keys or [])
        for mk in merge_keys:
            merge_vars = _find_variables_by_base(db, project_id, mk)
            for mv in merge_vars:
                if mv.id == keep_var.id:
                    continue
                for tid in mv.source_template_ids or []:
                    ids = list(keep_var.source_template_ids or [])
                    if tid not in ids:
                        ids.append(tid)
                        keep_var.source_template_ids = ids
                if mv.value and not keep_var.value:
                    keep_var.value = mv.value
                merged_from.append(mv.key)
                db.delete(mv)
                merged_count += 1

        if merged_from:
            keep_var.is_merged = True
            keep_var.merged_from_keys = merged_from

    db.commit()
    return {"merged_rows": merged_count}


def _find_primary_variable(db: Session, project_id: int, key: str) -> Variable | None:
    exact = _find_variable(db, project_id, key)
    if exact:
        return exact
    rows = _find_variables_by_base(db, project_id, key)
    return rows[0] if rows else None


def _find_variables_by_base(db: Session, project_id: int, base_key: str) -> list[Variable]:
    stmt = select(Variable).where(Variable.project_id == project_id)
    return [v for v in db.scalars(stmt).all() if _base_key(v.key) == base_key]


def _run_regex_validation(variables: list[dict[str, Any]]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    for var in variables:
        if not var.get("value"):
            if var.get("required"):
                issues.append(
                    {
                        "level": "error",
                        "variable_key": var["key"],
                        "message": f"必填变量 {var['key']} 未填写",
                        "suggestion": "请填写该变量",
                        "source": "regex",
                    }
                )
            continue
        err = validate_value(var["key"], var["value"], var.get("data_type", "text"))
        if err:
            issues.append(
                {
                    "level": "error",
                    "variable_key": var["key"],
                    "message": err,
                    "suggestion": "请修正格式后重试",
                    "source": "regex",
                }
            )
    return issues


async def ai_validate_variables(db: Session, project_id: int) -> dict[str, Any]:
    """基础正则校验 + AI 语义校验。"""
    variables = list_variables(db, project_id)
    regex_issues = _run_regex_validation(variables)

    ai_used = True
    ai_issues: list[dict[str, Any]] = []
    message: str | None = None
    try:
        filled = [
            {
                "key": v["key"],
                "label": v["label"],
                "value": v.get("value", ""),
                "category": v.get("category"),
                "data_type": v.get("data_type"),
                "required": v.get("required"),
            }
            for v in variables
            if v.get("value")
        ]
        result = await ai_service.validate_variable_data(filled, VALIDATION_RULES)
        ai_issues = result.get("issues", [])
        for issue in ai_issues:
            issue["source"] = "ai"
        issues = cross_validate_issues(ai_issues, regex_issues)
        _log_ai_call(
            db,
            project_id,
            ai_type="data_validate",
            prompt=str(filled),
            response=str(result),
            validation_result=result,
        )
    except ai_service.AIServiceUnavailableError as exc:
        ai_used = False
        message = str(exc)

    return {
        "regex_issues": regex_issues,
        "ai_issues": ai_issues,
        "issues": issues,
        "ai_used": ai_used,
        "message": message,
    }


def _log_ai_call(
    db: Session,
    project_id: int,
    *,
    ai_type: str,
    prompt: str,
    response: str,
    validation_result: dict | None = None,
) -> None:
    db.add(
        AILog(
            project_id=project_id,
            ai_type=ai_type,
            prompt=prompt[:4000],
            response=response[:4000],
            validation_result=validation_result,
        )
    )
    db.commit()


def _build_excel_workbook(rows: list[tuple[str, str, str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "变量"
    ws.append(list(EXCEL_HEADERS))
    for row in rows:
        ws.append(list(row))
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_template_excel(db: Session, project_id: int) -> bytes:
    """导出空白 Excel 模板。"""
    variables = list_variables(db, project_id)
    rows = [(v["key"], v["label"], "") for v in variables]
    return _build_excel_workbook(rows)


def export_variables_excel(db: Session, project_id: int) -> bytes:
    """导出已填写的变量数据。"""
    variables = list_variables(db, project_id)
    rows = [(v["key"], v["label"], v.get("value") or "") for v in variables]
    return _build_excel_workbook(rows)


def _parse_excel_rows(content: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header or len(header) < 3:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel 格式错误：需要三列（变量标识、变量名称、值）",
        )

    parsed: list[dict[str, Any]] = []
    for row_num, row in enumerate(rows_iter, start=2):
        if row is None or all(cell is None or str(cell).strip() == "" for cell in row[:3]):
            continue
        key = str(row[0]).strip() if row[0] is not None else ""
        value = str(row[2]).strip() if row[2] is not None else ""
        if not key:
            parsed.append({"row": row_num, "key": None, "value": value, "error": "缺少变量 key"})
            continue
        parsed.append({"row": row_num, "key": key, "value": value})
    return parsed


async def import_preview(
    db: Session,
    project_id: int,
    upload: UploadFile,
) -> dict[str, Any]:
    """Excel 导入预览（无状态，不写库）。"""
    get_project(db, project_id)
    content = await upload.read()
    try:
        parsed_rows = _parse_excel_rows(content)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"无法解析 Excel 文件: {exc}",
        ) from exc

    known = {v["key"]: v for v in list_variables(db, project_id)}
    success: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    for item in parsed_rows:
        if item.get("error"):
            errors.append(
                {
                    "row": item["row"],
                    "key": item.get("key"),
                    "message": item["error"],
                }
            )
            continue

        key = item["key"]
        value = item["value"]
        var = known.get(key)
        if var is None:
            errors.append(
                {
                    "row": item["row"],
                    "key": key,
                    "message": f"变量 {key} 不存在于当前项目",
                }
            )
            continue

        fmt_error = validate_value(key, value, var.get("data_type", "text"))
        if fmt_error:
            errors.append({"row": item["row"], "key": key, "message": fmt_error})
            continue

        success.append({"row": item["row"], "key": key, "value": value})

    total = len(parsed_rows)
    return {
        "success": success,
        "errors": errors,
        "summary": {
            "total": total,
            "succeeded": len(success),
            "failed": len(errors),
        },
    }


def import_variables(
    db: Session,
    project_id: int,
    rows: list[dict[str, Any]],
) -> dict[str, Any]:
    """确认导入变量（部分成功）。"""
    get_project(db, project_id)
    items = [{"key": r["key"], "value": r.get("value", "")} for r in rows if r.get("key")]
    results: dict[str, Any] = {
        "success": [],
        "errors": [],
        "summary": {"total": len(items), "succeeded": 0, "failed": 0},
    }

    any_written = False
    for index, item in enumerate(items, start=1):
        key = item["key"]
        value = item.get("value", "")
        var = _find_variable(db, project_id, key)
        if var is None:
            results["errors"].append(
                {"row": index, "key": key, "message": f"变量 {key} 不存在"}
            )
            continue
        fmt_error = validate_value(key, str(value), var.data_type)
        if fmt_error:
            results["errors"].append({"row": index, "key": key, "message": fmt_error})
            continue
        var.value = str(value)
        var.updated_at = datetime.now(UTC).replace(tzinfo=None)
        any_written = True
        results["success"].append({"key": key, "value": var.value})

    if any_written:
        db.commit()
        for row in results["success"]:
            var = _find_variable(db, project_id, row["key"])
            if var:
                row["updated_at"] = var.updated_at.isoformat()

    results["summary"]["succeeded"] = len(results["success"])
    results["summary"]["failed"] = len(results["errors"])
    return results
